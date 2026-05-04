"""Excel generator for WMW-738 Manager's Investigation form.

Builds the form programmatically — no external template file required.
"""
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ── Walmart brand colours ────────────────────────────────────────────────────
WMT_BLUE   = "0053E2"
WMT_SPARK  = "FFC220"
LIGHT_GREY = "F2F2F2"
MED_GREY   = "D9D9D9"
WHITE      = "FFFFFF"
DARK_TEXT  = "1A1A1A"

# ── Shared style helpers ─────────────────────────────────────────────────────
def _side(style="thin", color="000000"):
    return Side(style=style, color=color)


def _border(all_sides="thin"):
    s = _side(all_sides)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=11, color=DARK_TEXT, name="Calibri"):
    return Font(bold=bold, size=size, color=color, name=name)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _write(ws, cell_ref: str, value, bold=False, size=11,
           color=DARK_TEXT, bg=None, h_align="left", wrap=False):
    """Write a value to a cell with optional styling."""
    cell = ws[cell_ref]
    cell.value = value
    cell.font = _font(bold=bold, size=size, color=color)
    cell.alignment = _align(h=h_align, wrap=wrap)
    if bg:
        cell.fill = _fill(bg)


def _label(ws, cell_ref: str, text: str):
    """Write a form label (grey background, bold)."""
    _write(ws, cell_ref, text, bold=True, size=9,
           color=DARK_TEXT, bg=MED_GREY, h_align="right")


def _value(ws, cell_ref: str, text: str, wrap=False):
    """Write a form value (white background)."""
    _write(ws, cell_ref, text or "", size=11,
           color=DARK_TEXT, bg=WHITE, wrap=wrap)


def _merge_and_border(ws, cell_range: str, border_style="thin"):
    """Merge a range and apply a border to the merged cell."""
    ws.merge_cells(cell_range)
    b = _border(border_style)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = b


def _section_header(ws, cell_range: str, title: str):
    """Dark blue section header spanning a merged range."""
    ws.merge_cells(cell_range)
    top_left = ws[cell_range.split(":")[0]]
    top_left.value = title
    top_left.font = _font(bold=True, size=10, color=WHITE)
    top_left.fill = _fill(WMT_BLUE)
    top_left.alignment = _align(h="center")
    top_left.border = _border()


# ── Column widths ─────────────────────────────────────────────────────────────
COLUMN_WIDTHS = {
    "A": 14, "B": 10, "C": 12, "D": 8,  "E": 14,
    "F": 10, "G": 8,  "H": 8,  "I": 12, "J": 14,
    "K": 12, "L": 12, "M": 14, "N": 12, "O": 8, "P": 6,
}


def _setup_columns(ws):
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width


# ── Row-height helper ─────────────────────────────────────────────────────────
def _row_height(ws, row: int, height: float):
    ws.row_dimensions[row].height = height


# ─────────────────────────────────────────────────────────────────────────────
# PUBLIC API
# ─────────────────────────────────────────────────────────────────────────────
def generate_wmw738(data: dict) -> bytes:
    """Build a WMW-738 Manager's Investigation Excel workbook from *data*.

    Args:
        data: Dictionary containing all Phase 1 form fields.

    Returns:
        The workbook serialised as bytes, ready to write to disk or stream.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "WMW-738"

    _setup_columns(ws)

    def get(key, default=""):
        return data.get(key, default) or default

    # ── TITLE ROW ────────────────────────────────────────────────────────────
    _row_height(ws, 1, 28)
    ws.merge_cells("A1:P1")
    _write(ws, "A1",
           "WMW-738 — MANAGER'S INVESTIGATION REPORT",
           bold=True, size=14, color=WHITE, bg=WMT_BLUE, h_align="center")

    # ── SUBTITLE / SITE ROW (row 2) ──────────────────────────────────────────
    _row_height(ws, 2, 18)
    _label(ws,  "A2", "DC / FC #:")
    ws.merge_cells("B2:E2")
    _value(ws,  "B2", get("site_location"))
    _merge_and_border(ws, "B2:E2")

    _label(ws,  "F2", "TPR Code:")
    ws.merge_cells("G2:I2")
    _value(ws,  "G2", get("tpr_code"))
    _merge_and_border(ws, "G2:I2")

    _label(ws,  "J2", "Assoc. Manager:")
    ws.merge_cells("K2:P2")
    _value(ws,  "K2", get("associates_manager"))
    _merge_and_border(ws, "K2:P2")

    # ── SECTION: ASSOCIATE INFORMATION (rows 3–7) ────────────────────────────
    _row_height(ws, 3, 16)
    _section_header(ws, "A3:P3", "ASSOCIATE INFORMATION")

    _row_height(ws, 4, 18)
    _label(ws, "A4", "Last Name:")
    ws.merge_cells("B4:D4");  _value(ws, "B4", get("assoc_last_name"));  _merge_and_border(ws, "B4:D4")
    _label(ws, "E4", "First Name:")
    ws.merge_cells("F4:H4");  _value(ws, "F4", get("assoc_first_name")); _merge_and_border(ws, "F4:H4")
    _label(ws, "I4", "M.I.:")
    _value(ws, "J4", get("assoc_mi")); ws["J4"].border = _border()
    _label(ws, "K4", "Dept:")
    ws.merge_cells("L4:N4");  _value(ws, "L4", get("assoc_department")); _merge_and_border(ws, "L4:N4")
    _label(ws, "O4", "Schedule:")
    _value(ws, "P4", get("assoc_schedule")); ws["P4"].border = _border()

    _row_height(ws, 5, 18)
    _label(ws, "A5", "Hire Date:")
    ws.merge_cells("B5:D5");  _value(ws, "B5", get("assoc_hire_date")); _merge_and_border(ws, "B5:D5")
    _label(ws, "E5", "Norm. Duties?")
    _value(ws, "F5", get("doing_normal_duties")); ws["F5"].border = _border()
    _label(ws, "G5", "Norm. Shift?")
    _value(ws, "H5", get("on_normal_shift")); ws["H5"].border = _border()
    _label(ws, "I5", "Overtime?")
    _value(ws, "J5", get("on_overtime")); ws["J5"].border = _border()
    _label(ws, "K5", "If not normal:")
    ws.merge_cells("L5:P5");  _value(ws, "L5", get("if_not_normal_duties")); _merge_and_border(ws, "L5:P5")

    # ── SECTION: INCIDENT DETAILS (rows 6–11) ────────────────────────────────
    _row_height(ws, 6, 16)
    _section_header(ws, "A6:P6", "INCIDENT DETAILS")

    _row_height(ws, 7, 18)
    _label(ws, "A7", "Incident Date:")
    ws.merge_cells("B7:D7");  _value(ws, "B7", _fmt_date(get("incident_datetime"))); _merge_and_border(ws, "B7:D7")
    _label(ws, "E7", "Incident Time:")
    ws.merge_cells("F7:H7");  _value(ws, "F7", _fmt_time(get("incident_datetime"))); _merge_and_border(ws, "F7:H7")
    _label(ws, "I7", "Reported Date:")
    ws.merge_cells("J7:L7");  _value(ws, "J7", _fmt_date(get("reported_datetime"))); _merge_and_border(ws, "J7:L7")
    _label(ws, "M7", "Reported Time:")
    ws.merge_cells("N7:O7");  _value(ws, "N7", _fmt_time(get("reported_datetime"))); _merge_and_border(ws, "N7:O7")
    _label(ws, "P7", "Rpt'd To:")

    _row_height(ws, 8, 18)
    _label(ws, "A8", "Reported To:")
    ws.merge_cells("B8:E8");  _value(ws, "B8", get("reporting_manager")); _merge_and_border(ws, "B8:E8")
    _label(ws, "F8", "Location:")
    ws.merge_cells("G8:L8");  _value(ws, "G8", get("incident_location"));  _merge_and_border(ws, "G8:L8")
    _label(ws, "M8", "Incident Type(s):")
    incident_types = get("incident_types", [])
    if isinstance(incident_types, list):
        types_str = ", ".join(incident_types)
    else:
        types_str = str(incident_types)
    ws.merge_cells("N8:P8");  _value(ws, "N8", types_str); _merge_and_border(ws, "N8:P8")

    # ── SECTION: POWER EQUIPMENT (row 9) ─────────────────────────────────────
    _row_height(ws, 9, 16)
    _section_header(ws, "A9:P9", "POWER EQUIPMENT (if applicable)")

    _row_height(ws, 10, 18)
    _label(ws, "A10", "Make:")
    ws.merge_cells("B10:E10");  _value(ws, "B10", get("equipment_make"));     _merge_and_border(ws, "B10:E10")
    _label(ws, "F10", "Model:")
    ws.merge_cells("G10:J10");  _value(ws, "G10", get("equipment_model"));    _merge_and_border(ws, "G10:J10")
    _label(ws, "K10", "Asset ID:")
    ws.merge_cells("L10:P10");  _value(ws, "L10", get("equipment_asset_id")); _merge_and_border(ws, "L10:P10")

    # ── SECTION: ASSOCIATE STATEMENT (rows 11–30) ─────────────────────────────
    _row_height(ws, 11, 16)
    _section_header(ws, "A11:P11", "ASSOCIATE STATEMENT")

    def _textarea(ws, label_row, label_text, value_start_row, value_end_row,
                  field_value, row_height=60):
        _row_height(ws, label_row, 14)
        ws.merge_cells(f"A{label_row}:P{label_row}")
        _write(ws, f"A{label_row}", label_text, bold=True, size=9,
               bg=LIGHT_GREY, h_align="left")
        ws[f"A{label_row}"].border = _border()

        _row_height(ws, value_start_row, row_height)
        cell_range = f"A{value_start_row}:P{value_end_row}"
        ws.merge_cells(cell_range)
        _value(ws, f"A{value_start_row}", field_value, wrap=True)
        ws[f"A{value_start_row}"].alignment = _align(h="left", v="top", wrap=True)
        _merge_and_border(ws, cell_range)

    _textarea(ws, 12, "How did the incident occur? (Associate's account)",
              13, 15, get("how_incident_occurred"), row_height=70)
    _textarea(ws, 16, "Injury / illness description (body part, nature of injury):",
              17, 19, get("injury_description"), row_height=55)
    _textarea(ws, 20, "Objects / equipment / substances involved:",
              21, 23, get("objects_involved"), row_height=55)
    _textarea(ws, 24, "Safety accountability history:",
              25, 27, get("safety_accountability"), row_height=55)

    # ── SECTION: MANAGER ANALYSIS (rows 28–38) ────────────────────────────────
    _row_height(ws, 28, 16)
    _section_header(ws, "A28:P28", "MANAGER ANALYSIS")

    _textarea(ws, 29, "Unsafe behaviors / conditions contributing to incident:",
              30, 33, get("behavior_conditions"), row_height=70)
    _textarea(ws, 34, "How could this incident have been prevented?",
              35, 38, get("prevention"), row_height=70)

    # ── SECTION: WITNESSES (row 39–40) ───────────────────────────────────────
    _row_height(ws, 39, 16)
    _section_header(ws, "A39:P39", "WITNESSES")

    _row_height(ws, 40, 22)
    _label(ws, "A40", "Witness Name(s):")
    ws.merge_cells("B40:P40")
    _value(ws, "B40", get("witness_names"))
    _merge_and_border(ws, "B40:P40")

    # ── FOOTER ───────────────────────────────────────────────────────────────
    _row_height(ws, 41, 14)
    ws.merge_cells("A41:P41")
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    _write(ws, "A41",
           f"Generated by Incident Report Builder  •  {generated_at}",
           size=8, color="888888", bg=LIGHT_GREY, h_align="center")

    # ── PRINT SETTINGS ───────────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ── Date/time format helpers (module-private) ─────────────────────────────────
def _fmt_date(dt_str: str) -> str:
    if not dt_str:
        return ""
    try:
        return datetime.fromisoformat(dt_str).strftime("%m/%d/%Y")
    except (ValueError, TypeError):
        return dt_str


def _fmt_time(dt_str: str) -> str:
    if not dt_str:
        return ""
    try:
        return datetime.fromisoformat(dt_str).strftime("%I:%M %p")
    except (ValueError, TypeError):
        return dt_str
